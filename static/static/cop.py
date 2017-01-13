import xlrd

# fio = 'Богдан Дарья Александровна'
# month = 'сентябрь'
# date = 10

workbook = xlrd.open_workbook('export.xlsx')#open book by xlrd
sheet = workbook.sheet_by_index(0)# get sheet
# ============================================= finding row and col
for i in range(sheet.nrows):
    data = sheet.cell_value(i, 0)
    if data == fio:
        currow = i
        break
for i in range(sheet.ncols):
    data = sheet.cell_value(0, i)
    if data == month:
        curcol = is
        break
for i in range(curcol, sheet.ncols - curcol):
    data = sheet.cell_value(1, i)
    if data == date:
        curcol = i
        break
# ---------------------------------------------


from openpyxl import load_workbook
import warnings
warnings.simplefilter("ignore")# get rid of the warnings
book=load_workbook('export.xlsx')# open book by openpyxl
sheet = book.active # get sheet

sheet.cell(row=currow, column=curcol).value = "X" # write in the current cell
book.save('export.xlsx')# saving book
