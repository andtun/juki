# This Python file uses the following encoding: utf-8

#import time
from insertPoint import addPoint, convert, rdname
import os
import bottle
import requests
import xlrd
import beaker.middleware
import json
import insertPoint
import UserDB
import uuid
import sendEmail
from bottle import *
from funcslist import *
from passlib.hash import pbkdf2_sha256
from UserDB import db
from openpyxl import load_workbook
from datetime import datetime

def check_pw(user, pw):
    if user == "a":
        if pw == "b":
            return True
    return False

@route("/allDayNo")
def putAllNo():

    def allDayNo(filename):
        d = str(datetime.now())
        d = d[:d.find(" ")]
        d = d.split("-")
        month = convert(d[1])
        date = int(d[2])

        workbook = xlrd.open_workbook(filename)
        sheet = workbook.sheet_by_index(0)

        for i in range(sheet.ncols):
            data = sheet.cell_value(0, i)
            if data == month.decode("utf-8"):
                break

        curcol = i + date

        for i in range(3, sheet.nrows + 1):
            book = load_workbook(filename)
            sheet = book.active
            sheet.cell(row=i, column=curcol).value = "H"
            print(i, curcol)
            book.save(filename)

    fileList = ['8kl.xlsx', '9kl.xlsx', '10kl.xlsx', '11kl.xlsx']
    for i in fileList:
    	print("table filling for %s") % i
        allDayNo(i)
        print("!success!")


@hook('before_request')
def setup_request():
    #time.sleep(0.159)
    request.session = request.environ['beaker.session']
    #request.session.save()

    
#=====================USER PAGES========================#
    
@get("/")   #login html
def login():
    response.set_cookie("forgot", "not_yet")
    if not request.get_cookie("failed_login"):
        response.set_cookie("failed_login", 'undefined')

    if 'username' in request.session:
        if request.session['username'] != "":
            redirect("/menu")
            
    return stat_file("login.html")


@post("/")   #processing login info
def chklgn():
    postdata = request.body.read()
    print(postdata)
    cut = postdata.find("|")
    username = postdata[:cut]    #getting usrname & pw
    password = postdata[cut+1:]

    if check_login(username, password):    #if already in, you'll be redirected to the menu page
        #request.session['access'] = UserDB.get(request.session['username']).access_level     #setting atributes of the cookie
        response.set_cookie("failed_login", 'succeded')
        request.session['username'] = username
        print("EVENT:    user " + request.session['username'] + " logged in successfuly")
        redirect("/menu")

    else:   # if password and login don't match
        print("EVENT:   failed login")
        response.set_cookie("failed_login", 'failed')


@get("/menu")   # main page for the user
@need_auth      # need_auth decorator: if not authorized, you'll get 401 Error
def menu():

    if access_is('admin'):
        redirect('/main?adm')
    else:
        return stat_file('main.html')

    redirect('/logout')


#!!!REQUEST.SESSION['USERNAME'] WILL RETURN THE USRNAME OF THE LOGGED IN USER!!!
        

@get("/logerror")   # if pw didn't match login
def logerror():
    redirect("/")


@route("/main")   # main page
@need_auth
def main():
    return main_page(get_access())


#--------------------working with calendar-----------------------

@route("/submit", method="POST")
@need_auth
def do_form():
    print(request.body.read())
    return do_calendar_form(get_access())
#--------------------------------------------------------------------

@route("/fileDownload")
@need_auth
def download():
    filename = get_access() + ".xlsx"
    return static_file(filename, root='.', download=True)


@get("/logout")
def lout():
    logout()
    
    #redirect ("/")
    redirect("/main?1")


@get("/forgot_password")
def forgot():
    return stat_file("forgot_password.html")


@post("/forgot_password")
def forgot():
    username = request.body.read()
    #response.set_cookie("forgot", "not_yet")
    if UserDB.check(username):
        response.set_cookie("forgot", "OK")
        code = str(uuid.uuid4())
        email = UserDB.get(username).email
        print(email)
        send_message(email, code)
        UserDB.new_restore(username, code)
        return "Проверьте свою почту"
    else:
    	response.set_cookie("forgot", "failed")
    
    #return 'Done it for ', new_username, email

@get("/restore")
def restore_psw():
    code = request.query.code
    request.session['restore'] = code
    #response.set_cookie("test", "cookietest")
    if UserDB.check_link(code):
        print('check_link')
        return stat_file("restore_pswd.html")
    return "wrong link"

@post("/restore")
def restore_prcss():
    code = request.session['restore']
    print("CODE IS POST!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!! = "+code)
    username = request.forms.get('username')
    new_psw = request.forms.get('new_password')
    new_psw_conf = request.forms.get('new_password_repeated')

    print(code, username, new_psw, new_psw_conf)
    
    if UserDB.check_link(code):
        print('check_link=checked')
        if UserDB.check_code(username, code):
            print('check_code=checked')
            if new_psw == new_psw_conf:
                UserDB.set(username, 'pw', hsh(new_psw))
                cmnd = "DELETE FROM RestoreList WHERE username='%s';" % username
                UserDB.db.query(cmnd)
                response.set_cookie("forgot", "from_restore")
                redirect('/')
                return "done"
            else:
                response.set_cookie("restore", "Пароли не совпадают")
                redirect('/restore')
                return ("err pw") #psw dont match
            
    response.set_cookie("restore", "Что-то не так :(")
    redirect('/restore') #something wrong
            
            
@get("/new_psw_login") 
def itt(): 
	return stat_file('new_psw_login.html')

@get('/check_code')
def code():
	return request.session['restore']

@get("/change_password")    # change password html
@need_auth
def chngpsw_html():
    return stat_file("change_pswd.html")


@post("/change_password")   # processing the info from html, changing password
@need_auth
def chngpsw_process():
    
    its_username = request.session['username']
    old_password = request.forms.get('old_password')
    new_password = request.forms.get('new_password')
    new_password_conf = request.forms.get('new_password_repeated')
    
    if new_password != new_password_conf:
        return stat_file("smth_wrong_pwd.html")
    
    global d
    if (UserDB.check(its_username) and (pbkdf2_sha256.verify(old_password, UserDB.get(its_username).pw))):   # if the username exists and the old password is correct
        
        UserDB.set(its_username, 'pw', pbkdf2_sha256.hash(new_password, rounds=HRN))  # generate new pwd hash
        logout()
        
        return stat_file("pwd_changed.html")
    return stat_file("smth_wrong_pwd.html")


@get("/check_user")
@need_auth
def chk_usr():
    return UserDB.get(request.session['username']).fio.decode('unicode_escape')

#======================================================================
#                     ADMIN STUFF

@route("/showuserlist")
@need_auth
def showusr():
    if access_is('admin'):
        ulist = str(UserDB.db.fetch('SELECT username, fio, access_level FROM Userlist;')).decode('unicode_escape')
        return ulist


@route("/userlistdownload")
@need_auth
def downloadusr():
    if access_is('admin'):
        
        ulist = open('usrlist.txt', 'w')
        ulist.writelines(str(access))
        ulist.close()
        
    return static_file("usrlist.txt", root='.', download=True)

@route("/delete_user")
@need_auth
def delusr():
    if access_is('admin'):

        username = str(request.query.username)  #username is the one we entered
        
        if UserDB.check(username):   # deleting user
            UserDB.delete(username)
            
        return ("User "+username+" deleted!")
    else:
        return "No such user"


@post("/add_user")
@need_auth
def addusr():

    if access_is('admin'):
        
        his_username = request.forms.get('username')    #getting info from the form
        his_password = hsh(request.forms.get('password'))
        his_access_level = request.forms.get('access_level')
        his_email = request.forms.get('email')
        his_FIO = request.forms.get('FIO')
        
        if UserDB.check(his_username):
            return "User already exists"
        
        UserDB.add(his_username, his_password, his_FIO, his_access_level, his_email)

        return ("created user: username="+his_username+", password="+his_password+", access_level="+his_access_level+", FIO="+his_email)


@post("/change_access")
@need_auth
def chngaccs():

    if access_is('admin'):
        
        his_username = request.forms.get('username')
        new_access_level = request.forms.get('access_level')
        
        if UserDB.check(his_username):
            cmnd = "UPDATE UserList SET access_level='%s' WHERE username='%s';" % (new_access_level, his_username)
            db.query(cmnd)
            return ("Access level for "+his_username+" changed to "+ new_access_level)

        return "No such user"

@post("/change_email")
@need_auth
def chngemail():
    global d
    global email
    if access_is('admin'):
        
        his_username = request.forms.get('username')
        new_email = request.forms.get('new_mail')
        
        if UserDB.check(his_username):
            cmnd = "UPDATE UserList SET fio='%s' WHERE username='%s';" % (new_email, his_username)
            db.query(cmnd)
            return ("FIO for "+his_username+" changed to "+ new_email)

        return "No such user"



@get("/download_db")
@need_auth
def gt_accs():
    if access_is('admin'):
        filename = 'UserList.db'
        return static_file(filename, root='.', download = True)


@get("/setallno")
def alln():
    allNo()
    return "All No filled"


#=======================================================================
    #==================UPLOADING JSON========================
#=======================================================================
 

def postinfo():
    raw_body = request.body.read()
    print("raw body:")
    print(raw_body)
    new_events = request.files.get('sentJson')
    new_events = new_events.file.readlines()[0]
    new_events = new_events.decode("utf-8").encode("unicode_escape")
    print("new events: "+ new_events)


    dct = json.loads(new_events)

    print("DCT IS:", dct)

    print("-----  filling table from db started  -----")

    maildata = []
    
    for event in dct.keys():
        wrkdct = dct[event]
        print(wrkdct)
        name = wrkdct['personInfo']['last_name'].encode('utf-8') + ' ' + wrkdct['personInfo']['first_name'].encode('utf-8')
        form = wrkdct['personInfo']['class']
        print(name)
        a = {}
        a['inOrOut'] = wrkdct['inOrOut']
        a['name'] = name
        a['class'] = form
        maildata.append(a)
        
        #try:
        #    name = rdname(name)
        #except ValueError:
        #    print("Value Error in app.py")
        #    pass
        m = dct[event]['datetime']['date']
        month = m[m.find('.')+1:m.rfind('.')]
        day = m[:m.find('.')]
        # topr = name + " " + convert(month) + " " + str(day)
        # print(topr)
        addPoint(name, convert(month), int(day), form)

    print("-----  table filling finished successfully  -----")
    sendEmail.send(maildata)


@post("/post_info")
def pinfo():
    postinfo()

@put("/post_info")
def putinfo():
    postinfo()

#======================================================================
#======================================================================
#                       STYLES & IMAGES

@route("/<filename>")
def fileReturn(filename):
	return stat_file(filename)

@get("/dist/<p>/<fname>")
def dp(p,fname):
	return static_file(fname, root="./static/static/alco")
#======================================================================
#                        ERRORS CATCHING

@bottle.error(500)
def ff(error):
    print("!!!!!  something went wrong (500) error occurred  !!!!!")
    return stat_file("err500page.html")

@bottle.error(404)
def notfound(error):
    return stat_file("404error.html")

@bottle.error(401)
def fff(error):
    print("!unauthorized access try!")
    redirect("/")

@bottle.error(405)
def fff(error):
    print("!!!  error 405 (method not allowed)  detected  !!!")

#========================================================================
#===========================RUN========RUN===============================

bottle.run(app=app, host="0.0.0.0", port=os.environ.get('PORT', 5000), quiet=False)

