from datetime import datetime
import xlrd
from funcslist import convert
from openpyxl import load_workbook

def allDayNo(filename):
    d = str(datetime.now())
    d = d[:d.find(" ")]
    d = d.split("-")
    month = convert(d[1])
    date = int(d[2])

    workbook = xlrd.open_workbook(filename)
    sheet = workbook.sheet_by_index(0)

    for i in range(sheet.ncols):
        data = sheet.cell_value(0, i)
        if data == month.decode("utf-8"):
            break

    curcol = i + date

    for i in range(3, sheet.nrows + 1):
        book = load_workbook(filename)
        sheet = book.active
        sheet.cell(row=i, column=curcol).value = "H"
        book.save(filename)

print("starting all day n0")

fileList = ['8kl.xlsx', '9kl.xlsx', '10kl.xlsx', '11kl.xlsx']

for i in fileList:
    print("table filling for %s") % i
    allDayNo(i)
    print("!stage finished successfully!")