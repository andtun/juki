# This Python file uses the following encoding: utf-8

import os
import bottle
import requests
import xlrd
import beaker.middleware
from openpyxl import load_workbook
from passlib.hash import pbkdf2_sha256
from bottle import *
import UserDB
#from socket import gethostname, gethostbyname

import time



#================DECLARING CONSTANTS (NAMES USE ONLY CAPS)=================#

#session encryption key
RANDOMKEY = '''8RCURFLxYpQvehdTSe2I
K76Kl6d5V1A9MGaGggni
cfEjHrIG9tHHvkEeddkM
zvXYeN1tmK9PKRvwb8V5
cquNfumzsM5qEhkNEXsM'''


#encryption validate key
VALIDATEKEY = 'JUKI'


#cookie lifetime, seconds
CTIME = 1800 # 1/2 hour


#hashing round number (how many times we apply hash algorythm)
HRN = 20000


#static file root directory
STAT_FILE_ROOT = 'static/static/alco/'
MAIN_FILE_ROOT = 'static/static/classes/'




#=======================DESCRIBING STUFF=========================#

#bottle.debug(True)


session_opts = {
    'session.type': 'cookie',
    'session.data_dir': './session/',
    'session.auto': True,
    'session.cookie_expires': True,
    'session.encrypt_key': RANDOMKEY,
    'session.validate_key': VALIDATEKEY,
    'session.timeout': CTIME,
    #'session.type': 'cookie',
    #'session.type': 'file',
    #'session.validate_key': True,
    #'session.secure': True,
}

app = beaker.middleware.SessionMiddleware(bottle.app(), session_opts)


#------DICS FOR AUTH-------------------------------------------------------#

def hsh(password): #hashing pwd
    return pbkdf2_sha256.hash(password, rounds=HRN)


#--------------------------------------------------------------------------

import random, string

def randomword():
   return ''.join(random.choice(string.lowercase) for i in range(128))


#function returns True if login and pwd match
def check_login(username, password):
    if UserDB.check(username):
        print(UserDB.get(username).pw)
        return pbkdf2_sha256.verify(password, UserDB.get(username).pw)
    return False

#request.session['logged_in'] = False


import smtplib
from email.MIMEMultipart import MIMEMultipart
from email.MIMEText import MIMEText

def send_message(toaddr, code):
    fromaddr = "noreply.intschool@gmail.com"
    msg = MIMEMultipart()
    msg['From'] = fromaddr
    msg['To'] = toaddr
    msg['Subject'] = "Восстановление пароля"
    print(code)
    body = """Вы (или кто-то другой, выдающий себя за Вас) хотели восстановить пароль для своей учётной записи в системе контроля посещаемости школы 'Интеллектуал'.
Чтобы придумать новый пароль, перейдите по ссылке: https://int-school.herokuapp.com/restore?code=%s

----------------

Team JUKI""" % str(code)

    print("!!!!!!!!!!!!!!!!!!!!1EMAIL CODE IS = "+str(code))


    msg.attach(MIMEText(body, 'plain'))
     
    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.starttls()
    server.login(fromaddr, "adminpsw")
    text = msg.as_string()
    server.sendmail(fromaddr, toaddr, text)
    server.quit()



#===============DECORATORS&FUNCTIONS====================#

def need_auth(webpage):   #see how it works in the code down
    def wrapper():
        #request.session = request.environ['beaker.session']        
        if 'username' in request.session:
            #print("logged in !!!!!!!!")            
            if request.session['username'] != "":
                return webpage()
            
        redirect("/")

    return wrapper


def access_is(access_level):   #used to check the access level
    print(UserDB.get(request.session['username']).access_level)
    ans = (access_level == UserDB.get(request.session['username']).access_level)
    return ans

def get_access():
	return str(UserDB.get(request.session['username']).access_level)

def stat_file(filename):   #an easier way to return static file
    return static_file(filename, root = STAT_FILE_ROOT)


def main_page(filename):
	filename = str(filename) + ".html"
	return static_file(filename, root = MAIN_FILE_ROOT)


def logout():
    request.session['username'] = ""
    response.set_cookie("failed_login", 'undefined')



#=======================FORM METHODS==========================#    


def find_cell(fio, month, date, filename):        #find a cell
    
    fio = fio.decode("utf-8")
    fio = list(fio.split(" "))
    fio = fio[0] + " " + fio[1]

    print(fio, month, date, filename)

               #setting up table    
    workbook = xlrd.open_workbook(filename)
    sheet = workbook.sheet_by_index(0)
    curcol = 1
    currow = 1

               #finding fio
    for i in range(1, sheet.nrows):
        data = str(sheet.cell_value(i, 0).encode("unicode_escape"))
        print(data)
        data = list(data.split(" "))
        data = data[0] + " " + data[1]
        if data == fio:
            currow = i
            break

                #finding month
    for i in range(sheet.ncols):
        data = sheet.cell_value(0, i)
        if data == month.decode("utf-8"):
            curcol = i
            break

                #finding date
    curcol += int(date)

                #ending work
    book=load_workbook(filename)
    sheet = book.active
    currow += 1

    #return the cell we need and the table object, in which the cell is
    return currow, curcol, book, sheet



def do_calendar_form(form):     #filling the table using form (manual)

    
    def cal(calendar_str):  #turns yyyy-mm-dd  into  'monthname' and date
        indx = calendar_str.find('.')
        day = calendar_str[:indx]
        calendar_str = calendar_str[(indx+1):]
        indx = calendar_str.find('.')
        monthnum = calendar_str[:indx]
        
        month = {'01': "январь", '02': "февраль", '03': "март", '04': "апрель", '05': "май", '06': "июнь", '07': "июль",
             '08': "август", '09': "сентябрь", '10': "октябрь", '11': "ноябрь", '12': "декабрь"}
        
        return day, month[monthnum]


    filename = str(form) + ".xlsx"
           #getting info from the form
    data = request.body.read()
    cut = data.find("|")
    fio = data[:cut]
    data = data[cut+1:]
    cut = data.find("|")
    cal_str = data[:cut]
    YesNo = data[cut+1:]
    print(fio, cal_str, YesNo)
    
           #formatting info
    date, month = cal(cal_str)

            #finding the cell
    currow, curcol, book, sheet = find_cell(fio, month, date, filename)

    print(fio, month, date, filename)


            #deciding what to insert into the cell
    if YesNo == 'Yes':
        sheet.cell(row=currow, column=curcol).value = ""
    else:
        sheet.cell(row=currow, column=curcol).value = "H"
    book.save(filename)

    print("FORM FILLED SUCCESFULLY FOR^ ",filename,fio )

            #returning success page
    return stat_file("back.html")



def do_calendar_from_db(name, month, date):   #filling the table using DB (automatic)
    currow, curcol = find_cell(name, month, date)
    data = sheet.cell_value(currow, curcol)
    
    if data != X:  #if data is X, don't change the cell. If not X - the student is in school, so set the cell empty
        sheet.cell(row=currow, column=curcol).value = ""


#======================================USERDB HERE=================================================
#==============================================================================================

# no userdb
